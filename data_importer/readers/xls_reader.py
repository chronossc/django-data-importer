# coding: utf-8
import xlrd
import openpyxl
from .base import BaseReader


class XLSReader(BaseReader):
    def __init__(self,f,**kwargs):
        self._sheet_name = kwargs.pop('sheet',None)
        self._on_demand = kwargs.pop('on_demand',True)
        super(XLSReader,self).__init__(f)

    def set_reader(self):
        self.__workbook = xlrd.open_workbook(self._source.name,on_demand=self._on_demand)
        if self._sheet_name:
            self._reader = self.__workbook.sheet_by_name(self._sheet_name)
        else:
            self._reader = self.__workbook.sheet_by_index(0)

        self.nrows = self._reader.nrows
        self.ncols = self._reader.ncols

    @property
    def headers(self):
        if not self._headers:
            self._headers = map(self.normalize_string,[self._reader.cell(0,c).value for c in range(self.ncols)])
        return self._headers

    def get_items(self):
        for r in range(1,self.nrows):
            values = [self._reader.cell(r,c).value for c in range(self.ncols)]
            if not all(values): continue
            yield self.get_item(values)

class XLSXReader(XLSReader):

    def set_reader(self):
        self.__workbook = openpyxl.reader.excel.load_workbook(self._source)
        if self._sheet_name:
            self._reader = self.__workbook.worksheets[self.__workbook.get_sheet_names().index(self._sheet_name)]
        else:
            self._reader = self.__workbook.worksheets[0]
    
    @property
    def headers(self):
        if not self._headers:
            self._headers = map(self.normalize_string,[c.value for c in self._reader.rows[0]])
        return self._headers

    def get_items(self):
        for row in self._reader.rows[1:]:
            values = [c.value for c in list(row)]
            if not all(values): continue
            yield self.get_item(values)


